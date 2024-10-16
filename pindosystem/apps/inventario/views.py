from django.shortcuts import render, redirect




from django.db import transaction

from rodales.models import Rodales
from rodales_gis.models import Rodalesgis
from django.contrib import messages
from login.models import Users
from emsefor.models import Emsefor

from inventario.models import InventariosCategories, InventariosObservaciones, InventariosDamages, InventariosRelevamientos, TYPES_RELEVAMIENTO, \
ArbolesRelevamientoPoda, InventariosParcelasgis, ArbolesRelevamientoOthers, ResumenRelevamientos
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponseRedirect, HttpResponse
from django.db import IntegrityError
from django.urls import reverse

from plantaciones.models import Plantaciones

from intervenciones.models import Intervenciones, INTERVENCIONES_DICT, PodaIntervencion

from configuration.models import IntervencionesTypes


from inventario.utility import calculate_general_resumen_poda, process_excel_tradicional, process_excel_poda, calculate_resumen_poda, calculate_resumen_others

from django.db.models import Sum, F, Count, Q, Avg, Value, Max, Func, FloatField, ExpressionWrapper
from django.db.models.functions import Concat
from django.db.models.functions.comparison import NullIf
from django.core.serializers import serialize
import json

import openpyxl




    #tipos de parcelas
types_parcels = (
       ( 'Rectangular', 'Rectangular' ),
       ( 'Triangular', 'Triangular' ),
       ( 'Circular', 'Circular' ),
       ( 'Otro', 'Otro' ),
)


# Create your views here.

@login_required
def indexInventariosRelevamientos(request):
    
    relevamientos = InventariosRelevamientos.objects.all() \
        .values(
        'pk', 'number', 'type', 'h_deseada', 'h_last_poda', 'fecha', 'created', 
        'is_finish', 'user__first_name', 'user__last_name', 'user_relevador', 'user_relevador__first_name', 'user_relevador__last_name',
        'rodales_id', 'rodales__cod_sap', 'categorias_id', 'categorias__name', 'emsefor__name', 'invrel_resumen__tam_parcela'
        ).annotate(cant_parcelas = Count('parc_invrel')) \
        .annotate(cant_arboles = Sum('parc_invrel__number_trees'), cant_res = Count('invrel_resumen', filter=Q(type = 'Preexistente'))).order_by('number')
    
  


    context = {'relevamientos' : relevamientos, 
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Administración de los Relevamientos de Inventarios'}

    return render(request, 'configuration/inventarios/relevamientos/index.html', context)



#traigo las elecciones

@login_required
def addInventarioRelevamientos(request):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Crea un Relevamiento de Inventario'}
    

     #traigo los rodales que estan relacionados con este sagpya
    rodales = Rodales.objects.exclude(cod_sap = None).values_list("rodales_id", "cod_sap")
       
    context.update({'rodales' : rodales})


    categorias = InventariosCategories.objects.all().values_list('pk', 'name')
    context.update({'categorias' : categorias})


    emsefors = Emsefor.objects.all().values_list('pk', 'name').order_by('name')
    context.update({'emsefors' : emsefors})

    #envio los usarios tmb
    users = Users.objects.annotate(fullname = Concat('last_name', Value(' '), 'first_name' ))\
        .values_list('pk', 'fullname').order_by('fullname')
    
    context.update({'types_relevamiento' : TYPES_RELEVAMIENTO})
    
    context.update({'responsable_user' : users})

    if request.method == 'POST':
        try:

            #creo que tambien el resumen

            with transaction.atomic():

        
                type = request.POST.get('type_rel')
                h_deseada = request.POST.get('h_deseada')
                h_last_poda = request.POST.get('h_last_poda')
                rodal_id = request.POST.get('select-rodal')

                fecha = request.POST.get('fecha')
                categoria = request.POST.get('select-categoria')

                responsable = request.POST.get('select-responsable')

                user_responsable = Users.objects.get(pk=responsable)
        
                user_entity = Users.objects.get(pk=request.user.pk)
                rodal = Rodales.objects.get(pk=rodal_id)

                #select-emsefor
                emsefor_id = request.POST.get('select-emsefor')

                is_finish = True if request.POST.get('is_finish') == 'on' else False 
            

                number = len(InventariosRelevamientos.objects.filter(rodales = rodal).values()) + 1
    

                #creo el objeto
                result = InventariosRelevamientos.objects.create(number = number, type = type,
                                                                    h_deseada = h_deseada, 
                                                                    h_last_poda = h_last_poda,
                                                                    user = user_entity,
                                                                    fecha = fecha,
                                                                    rodales = rodal, categorias_id = categoria,
                                                                    user_relevador = user_responsable, 
                                                                    is_finish = is_finish, emsefor_id = emsefor_id)
            

                #creo el resumen vacio
                res_rel = ResumenRelevamientos.objects.create(relevamientos = result, user = user_entity)


                messages.success(request, "El Relevamiento ha sido agregado con éxito!.")
                return redirect('inventarios-relevamientos-index')
    
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'configuration/inventarios/relevamientos/add.html', context)
  

    else:
    
        return render(request, 'configuration/inventarios/relevamientos/add.html', context)


@login_required
def editInventarioRelevamientos(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Editar un Relevamiento de Inventario'}
    
    try:
        #traigo la entidad
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

        context.update({'relevamiento' : relevamiento})

        
        rodales = Rodales.objects.exclude(cod_sap = None).values_list("rodales_id", "cod_sap")
       
        context.update({'rodales' : rodales})


        categorias = InventariosCategories.objects.all().values_list('pk', 'name')
        context.update({'categorias' : categorias})

        emsefors = Emsefor.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'emsefors' : emsefors})

        #envio los usarios tmb
        users = Users.objects.annotate(fullname = Concat('last_name', Value(' '), 'first_name' ))\
            .values_list('pk', 'fullname').order_by('fullname')
        
        context.update({'responsable_user' : users})

        if request.method == 'POST':
            
            h_deseada = request.POST.get('h_deseada')
            h_last_poda = request.POST.get('h_last_poda')
            rodal_id = request.POST.get('select-rodal')

            fecha = request.POST.get('fecha')
            categoria = request.POST.get('select-categoria')

            responsable = request.POST.get('select-responsable')

            user_responsable = Users.objects.get(pk=responsable)
       
            user_entity = Users.objects.get(pk=request.user.pk)
            rodal = Rodales.objects.get(pk=rodal_id)

            is_finish = True if request.POST.get('is_finish') == 'SI' else False

              #select-emsefor
            emsefor_id = request.POST.get('select-emsefor')

            #traigo la entidad
            catego_aux = InventariosCategories.objects.get(pk = categoria)

            #si la categoria es poda no se puede cambiar a otro y viceversa
            if relevamiento.categorias.name == 'Poda' and catego_aux.name != 'Poda': 
                messages.error(request, str('No se puede cambiar un Relevamiento de Poda por otra categoria'))
                return HttpResponseRedirect(reverse("inventarios-relevamientos-edit", args=[idrelevamiento]))
            
            
            elif relevamiento.categorias.name != 'Poda' and catego_aux.name == 'Poda':
                messages.error(request, str('No se puede cambiar un Relevamiento de ' + relevamiento.categorias.name + ' por la categoria de Poda' ))
                return HttpResponseRedirect(reverse("inventarios-relevamientos-edit", args=[idrelevamiento]))
            else:

                relevamiento.h_deseada = h_deseada
                relevamiento.h_last_poda = h_last_poda
                relevamiento.rodales = rodal
                relevamiento.fecha = fecha
                relevamiento.categorias_id = categoria
                relevamiento.user_relevador = user_responsable
                relevamiento.user = user_entity
                relevamiento.is_finish = is_finish
                relevamiento.emsefor_id = emsefor_id

                try:
                    relevamiento.save()
                    messages.success(request, "El relevamiento se ha editado con éxito")

                    return redirect('inventarios-relevamientos-index')

                except Exception as e:
                    messages.error(request, str(e))
                    return render(request, 'configuration/inventarios/relevamientos/edit.html', context)



    except InventariosRelevamientos.DoesNotExist as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')

    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')
    
    return render(request, 'configuration/inventarios/relevamientos/edit.html', context)
    





@login_required
def addArbolesRelevamiento(request, idrelevamiento):
    #si hay arboles cargados, muestro los arboles, sino
    #consulto por el tipo de relevamiento, si es poda tengo que ir a la tabla poda
    relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

    has_parcelas = False
    
    has_arboles = None

    #puede haber arboles -> seguro tiene parcelas
    #puede haber parcelas sin arboles
    #puede haber parcelas con arboles
    #si hay parcelas sin arboles, ingreso la cantidad de arboles por parcela

    if(relevamiento.categorias):

        if relevamiento.categorias.name == 'Poda':
            
            has_arboles = True if len(ArbolesRelevamientoPoda.objects.filter(relevamientos_id = idrelevamiento).values()) > 0 else False

            has_parcelas = True if len(InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento).values()) > 0 else False

            #si hay parcelas paso a cargar los arboles
            if has_parcelas:
                #consuto si hay arboles

                if has_arboles:
                    return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
                else:
                    #cargo los arboles pero paso las podas
                    suma_parc = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)\
                    .aggregate(suma = Sum('number_trees'))
                   
                    #controlar que no sea la suma 0
                    if int(suma_parc['suma']) > 0:

                        return HttpResponseRedirect(reverse("inventarios-relevamientos-addtrees-poda", args=[idrelevamiento, suma_parc['suma']]))
                    else:
                        #suma era igual a 0, llevamos a index
                        messages.error(request, str(e))
                        return HttpResponseRedirect(reverse("inventarios-relevamientos-index"))
                   
                       
            else:

                return HttpResponseRedirect(reverse("inventarios-relevamientos-numparc", args=[idrelevamiento]))

        else:
            #otra categoria que no es poda
            has_arboles = True if len(ArbolesRelevamientoOthers.objects.filter(relevamientos_id = idrelevamiento).values()) > 0 else False

            has_parcelas = True if len(InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento).values()) > 0 else False

            if has_parcelas:


                if has_arboles:
                    return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
                else:
                    #cargo los arboles
                    #tengo que sumar la cantidad de arboles de las parcelas
                    suma_parc = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)\
                    .aggregate(suma = Sum('number_trees'))
                   
                    #controlar que no sea la suma 0
                    if int(suma_parc['suma']) > 0:

                        #cargo los arboles
                        return HttpResponseRedirect(reverse("inventarios-relevamientos-addtrees-others", args=[idrelevamiento, suma_parc['suma']]))
                    
                    else:
                        #como era 0 tengo que enviar a cargar las parcelas

                        return HttpResponseRedirect(reverse("inventarios-relevamientos-numparc", args=[idrelevamiento]))
                    

            else:
                return HttpResponseRedirect(reverse("inventarios-relevamientos-numparc", args=[idrelevamiento]))

    else:
        return HttpResponseRedirect(reverse("inventarios-relevamientos-numparc", args=[idrelevamiento]))
    


@login_required
def enterNumParcelas(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Crea Parcelas de un Relevamiento de Inventario'}
    
    context.update({'idrelevamiento' : idrelevamiento})

    

    if request.method == 'POST':

        try:
        
      
            number = request.POST.get('number')

            if int(number) > 0:
                #tengo que ir a cargar el detalle de las parcelas
                #si hay parcelas hago e salto a arboles

                return HttpResponseRedirect(reverse("inventarios-relevamientos-parcelas-details", args=[idrelevamiento, number]))

           
            return redirect('inventarios-relevamientos-index')
    
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'configuration/inventarios/relevamientos/add_number_parcels.html', context)
  

    else:
    
        return render(request, 'configuration/inventarios/relevamientos/add_number_parcels.html', context)


@login_required
def addParcelsDetails(request, idrelevamiento, number_parcels):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    context.update({'idrelevamiento' : idrelevamiento,
                    'number_parcels' : number_parcels})
    
    #convierto a una lista los numero de arboles
    num_parcels = [(num + 1) for num in range(number_parcels)]
    
    context.update({'number_parcels_list' : num_parcels})



    context.update({'types_parcels' : types_parcels})


    if request.method == 'POST':
        number_parcel = request.POST.getlist('number_parcel')
        size_parcela = request.POST.getlist('size_parcela')
        type_parcel = request.POST.getlist('tipo')

        number_trees = request.POST.getlist('number_trees')

        #traigo el relevamiento
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)


        #se puede generar un INDEXERROR

        if relevamiento:

            lista_objetos = []
            num_total_trees = 0
            for n in range(number_parcels):

                lista_objetos.append(InventariosParcelasgis(size_parcela = size_parcela[n], number_parcela = num_parcels[n], 
                                                            type = type_parcel[n], rodales = relevamiento.rodales,
                                                            relevamiento = relevamiento, number_trees = number_trees[n]))
                num_total_trees =  num_total_trees + int(number_trees[n])
                    
            InventariosParcelasgis.objects.bulk_create(lista_objetos)

            #me voy a cargar los arboles
            #inventarios-relevamientos-addtrees-poda
            #consulto por el tipo de relevamiento para redireccionar

            if relevamiento.categorias.name == 'Poda':

                return HttpResponseRedirect(reverse("inventarios-relevamientos-addtrees-poda", args=[idrelevamiento, num_total_trees]))
            
            else:
                #otro tipo que no es poda
                return HttpResponseRedirect(reverse("inventarios-relevamientos-addtrees-others", args=[idrelevamiento, num_total_trees]))

       
        else:
            #sin relevamiento nada se puede hacer
            pass
    

    return render(request, 'configuration/inventarios/relevamientos/add_parcels_details.html', context)
    


@login_required
def enterNumArboles(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Crea un Relevamiento de Inventario'}
    

     #traigo los rodales que estan relacionados con este sagpya
 

    context.update({'idrelevamiento' : idrelevamiento})

    if request.method == 'POST':
        try:
        
      
            number = request.POST.get('number')

            if int(number) > 0:
                return HttpResponseRedirect(reverse("inventarios-relevamientos-addtrees-poda", args=[idrelevamiento, number]))       
           
            return redirect('inventarios-relevamientos-index')
    
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'configuration/inventarios/relevamientos/add_number_trees.html', context)
  

    else:
    
        return render(request, 'configuration/inventarios/relevamientos/add_number_trees.html', context)

@login_required
def addArbolesToInventarioRelevamientos(request, idrelevamiento, number_arboles):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    context.update({'idrelevamiento' : idrelevamiento,
                    'number_arboles' : number_arboles})
    

    try:


        #traigo las parcelas de este relevamiento
        parcelas = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)
        context.update({'parcelas' : parcelas})

        #convierto a una lista los numero de arboles
        num_arboles = [(num + 1) for num in range(number_arboles)]
        
        #context.update({'number_arboles_list' : num_arboles})

        #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})

        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        #reviso la cantidad de arboles por parcelas

        if request.method == 'POST':
            parcela_number = request.POST.getlist('parcela')
            numbers = request.POST.getlist('number_tree')
            dap = request.POST.getlist('dap')
            dmsm = request.POST.getlist('dmsm')
            h_poda = request.POST.getlist('h_poda')
            h_total = request.POST.getlist('h_total')
            is_poda = request.POST.getlist('is_poda')
            damages = request.POST.getlist('select-damage')
            observacion = request.POST.getlist('select-observacion')



            lista_objetos = []
            for n in range(number_arboles):

                #controlar que pueda dejar vacio el formulario de carga
                dap_ = dap[n] if dap[n] != '' else None
                dmsm_ = dmsm[n] if dmsm[n] != '' else None
                h_poda_ = h_poda[n] if h_poda[n] != '' else None
                h_total_ = h_total[n] if h_total[n] != '' else None
                
                lista_objetos.append(ArbolesRelevamientoPoda(number_tree = numbers[n], dap = dap_, dmsm = dmsm_, 
                                                             h_poda = h_poda_, h_total = h_total_, 
                                                is_poda = is_poda[n], damages_id = damages[n], 
                                                parcela_id = parcela_number[n], 
                                                observaciones_id = observacion[n],
                                                relevamientos_id = idrelevamiento, ))

            
            
            with transaction.atomic():
            
                ArbolesRelevamientoPoda.objects.bulk_create(lista_objetos)

                #agregue los arboes, entonces cargo las datas de resumen

                result = calculate_resumen_poda(idrelevamiento)

                if result == False:
                    raise





            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    except Exception as e:
        messages.error(request, str(e))
        return render(request, 'configuration/inventarios/relevamientos/add_trees_poda.html', context)


    return render(request, 'configuration/inventarios/relevamientos/add_trees_poda.html', context)



@login_required
def addArbolesToInventarioRelevamientosOthers(request, idrelevamiento, number_arboles):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    context.update({'idrelevamiento' : idrelevamiento,
                    'number_arboles' : number_arboles})
    

    try:
        #traigo las parcelas de este relevamiento
        parcelas = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)
        context.update({'parcelas' : parcelas})

        #convierto a una lista los numero de arboles
        num_arboles = [(num + 1) for num in range(number_arboles)]
        
        #context.update({'number_arboles_list' : num_arboles})

        #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})
                
        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        #reviso la cantidad de arboles por parcelas

        if request.method == 'POST':
            parcela_number = request.POST.getlist('parcela')
            numbers = request.POST.getlist('number_tree')
            dap = request.POST.getlist('dap')
            h_poda = request.POST.getlist('h_poda')
            h_total = request.POST.getlist('h_total')
            damages = request.POST.getlist('select-damage')
            observacion = request.POST.getlist('select-observacion')



            lista_objetos = []
            for n in range(number_arboles):

                dap_ = dap[n] if dap[n] != '' else None
                h_poda_ = h_poda[n] if h_poda[n] != '' else None
                h_total_ = h_total[n] if h_total[n] != '' else None
                
                lista_objetos.append(ArbolesRelevamientoOthers(number_tree = numbers[n], 
                                                               dap = dap_,  
                                                               h_poda = h_poda_, 
                                                               h_total = h_total_, 
                                                damages_id = damages[n], parcela_id = parcela_number[n], 
                                                observaciones_id = observacion[n],
                                                relevamientos_id = idrelevamiento))

            
            with transaction.atomic():
            
                ArbolesRelevamientoOthers.objects.bulk_create(lista_objetos)

                #agregue los arboes, entonces cargo las datas de resumen

                result = calculate_resumen_others(idrelevamiento)

                if result == False:
                    raise



            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))


        return render(request, 'configuration/inventarios/relevamientos/add_trees_others.html', context)

    except Exception as e:
         messages.error(request, str(e))
         return redirect('inventarios-relevamientos-index')

@login_required
def viewArbolesRelevamiento(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)
        context.update({'relevamiento' : relevamiento})



        parcelas = relevamiento.parc_invrel.all()
      
        context.update({'parcelas' : parcelas})
        
        if relevamiento.categorias:
            #no esta vacio entonces proceso
            if relevamiento.categorias.name == 'Poda':
                
                
                arboles = relevamiento.invrel_arb_poda.all()
                context.update({'arboles' : arboles})
                context.update({'cant_arboles' : len(arboles)})
   
                return render(request, 'configuration/inventarios/relevamientos/view_arboles.html', context)


            else:
                arboles = relevamiento.invrel_arb_others.all()
                context.update({'arboles' : arboles})
                context.update({'cant_arboles' : len(arboles)})

            
                
                return render(request, 'configuration/inventarios/relevamientos/view_arboles_others.html', context)



    except Exception as e:
         messages.error(request, str(e))
         return redirect('inventarios-relevamientos-index')


@login_required
def viewStatsRelevamiento(request, idrelevamiento):

    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)
        context.update({'relevamiento' : relevamiento})

        plantacion = Plantaciones.objects.get(rodales_id = relevamiento.rodales_id)

        context.update({'plantacion' : plantacion})

        rodales_gis = Rodalesgis.objects.get(rodales_id = relevamiento.rodales_id)
        
        context.update({'rodales_gis' : rodales_gis})

        intensity = 0
        #(tam_parc * cantidad_parcelas) / (sup_rodap * 1000)

        if relevamiento.categorias:
            #no esta vacio entonces proceso
            if relevamiento.categorias.name == 'Poda':
                
                
                arboles = relevamiento.invrel_arb_poda.all()
                context.update({'arboles' : arboles})
                context.update({'cant_arboles' : len(arboles)})

                #cargo los datos de damages
                damages = InventariosDamages.objects.all()
                context.update({'damages' : damages})

                observaciones = InventariosObservaciones.objects.all()
                context.update({'observaciones' : observaciones})

               
                #esto dejo porque es el resumen de cada parcela
                data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento) \
                .values('pk') \
                .annotate(
                            n_arb_ha = Count('invrel_parc_arb') * 10000 / F('size_parcela'), #CALCULAR NUMERO DE ARB POR HA
                            arboles_podados = Count('invrel_parc_arb', 
                                                    filter=Q(invrel_parc_arb__is_poda = True)) * 10000 / F('size_parcela'),
                            arb_h_deseada = Count('invrel_parc_arb', 
                                                  filter=Q(invrel_parc_arb__h_poda__gte = relevamiento.h_deseada)) * 10000 / F('size_parcela'), 
                            arb_no_podados = Count('invrel_parc_arb',  filter=Q(invrel_parc_arb__is_poda = False)) * 10000 / F('size_parcela'),
                            dap_ = Avg('invrel_parc_arb__dap'),
                            h_total_ = Avg('invrel_parc_arb__h_total', filter=Q( invrel_parc_arb__h_total__gt = Value('0'))),
                            h_poda_ = Avg('invrel_parc_arb__h_poda', filter=Q( invrel_parc_arb__h_poda__gt = Value('0'))),
                            dmsm_ = Avg('invrel_parc_arb__dmsm', filter=Q( invrel_parc_arb__dmsm__gt = Value('0'))),
                            area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb__dap') * F('invrel_parc_arb__dap'))) / 4) / 10000) 
                            / F('size_parcela'),
                            porc_remo = ExpressionWrapper(((F('h_poda_') - relevamiento.h_last_poda) / 
                                          NullIf(F('h_total_') - relevamiento.h_last_poda, 0)) * 100, 
                                          output_field=FloatField()),
                            tam_parc =  Avg('size_parcela')
                            ).values()
                
                context.update({'data_resumen' : data_resumen})

           
                #calculo lo mismo de arriba pero sin resumir
                res = ResumenRelevamientos.objects.get(relevamientos = relevamiento)
  

                #no lo calculo, sino que lo traigo desde la tabla de resumen
                #res = calculate_general_resumen_poda(data_resumen=data_resumen)
               
                context.update({'resumen_general' : res})

               

              
               


                
                '''  arb_h_deseada = Count('invrel_parc_arb', filter=Q(invrel_parc_arb__h_total__gte = Value(relevamiento.h_deseada), 
                                                                 invrel_parc_arb__h_total__lt = (relevamiento.h_deseada + 3))) * 10000 / F('size_parcela'), 
                         '''
             
                damages = InventariosDamages.objects.all().order_by('pk')
                context.update({'damages' : damages})

                
                
                #data de los damages
                data_damages = ArbolesRelevamientoPoda.objects.filter(relevamientos_id = idrelevamiento) \
                 .values('parcela__number_parcela', 'damages_id') \
                 .annotate(res_damage = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                 .values('parcela__number_parcela', 'damages_id', 'damages__name', 'parcela__size_parcela', 'res_damage')\
                 .order_by('parcela__number_parcela')
                
                               
                
                
                #traigo las parcelas
                parcelas = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)

                #rearmo los damages
                data_damages_new = []
                
                for parc in parcelas:
                   
                    data_damages_present = []
                    #cargo los da;os presentes
                    #recorro los damages completos

                    for damag_ in damages:
                        present = False
                        #este tiene el dato agreado traido de la tabla
                        for dam_data in data_damages:

                            #primero consulto si estoy en la parcela

                            if parc.number_parcela == dam_data['parcela__number_parcela']:

                                  if dam_data['damages_id'] == damag_.pk:

                                    present = True
                                    data_damages_present.append({dam_data['damages_id'] : dam_data['res_damage']})
                        
                        #no encontro ese damages, entonces lo cargo vacio
                        if present == False:
                            data_damages_present.append({damag_.pk : ''})
                    
                    #cargo la parcela tmb
                    data_damages_new.append({
                        'parcela' : parc.number_parcela,
                        'damages' : data_damages_present
                    })

              
                context.update({'data_damages' : data_damages_new})
                
               

                observaciones = InventariosObservaciones.objects.all().order_by('pk')
                context.update({'observaciones' : observaciones})
                 #data de los damages

                data_calidad = ArbolesRelevamientoPoda.objects.filter(relevamientos_id = idrelevamiento) \
                 .values('parcela__number_parcela', 'observaciones_id') \
                 .annotate(res_calidad = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                 .values('parcela__number_parcela', 'observaciones_id', 'observaciones__name', 'parcela__size_parcela', 'res_calidad')\
                 .order_by('parcela__number_parcela')
                
             
                data_calidad_new = []

               


                for parc in parcelas:
                   
                    data_obs_present = []
                    #cargo los da;os presentes
                    #recorro los damages completos

                    for obs_ in observaciones:
                        present = False
                        #este tiene el dato agreado traido de la tabla
                        for cal_data in data_calidad:

                            #primero consulto si estoy en la parcela

                            if parc.number_parcela == cal_data['parcela__number_parcela']:

                                  if cal_data['observaciones_id'] == obs_.pk:

                                    present = True
                                    data_obs_present.append({cal_data['observaciones_id'] : cal_data['res_calidad']})
                        
                        #no encontro ese damages, entonces lo cargo vacio
                        if present == False:
                            data_obs_present.append({obs_.pk : ''})
                    
                    #cargo la parcela tmb
                    data_calidad_new.append({
                        'parcela' : parc.number_parcela,
                        'observaciones' : data_obs_present
                    })

              
                context.update({'data_observaciones' : data_calidad_new})
                
                

               
              

                return render(request, 'configuration/inventarios/relevamientos/view_statistics.html', context)


            else:
                #aca traigo los datos para others arboles
                arboles = relevamiento.invrel_arb_others.all()
                context.update({'arboles' : arboles})
                context.update({'cant_arboles' : len(arboles)})

                 #cargo los datos de damages
                damages = InventariosDamages.objects.all()
                context.update({'damages' : damages})

                observaciones = InventariosObservaciones.objects.all()
                context.update({'observaciones' : observaciones})

                #CUANDO VIENEN VALORES VACIOS NO CONSIDERAR EN LOS CALCULOS DE LOS ARBOELS

              
                data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento) \
                .values('pk') \
                .annotate(
                            n_arb_ha = Count('invrel_parc_arb_others'),
                            arb_h_deseada = Count('invrel_parc_arb_others', 
                                                  filter=Q(invrel_parc_arb_others__h_total__gte = relevamiento.h_deseada, 
                                                                 invrel_parc_arb_others__h_total__lt = (relevamiento.h_deseada + 3))), 
                            dap_ = Avg('invrel_parc_arb_others__dap'),
                            h_total_ = Avg('invrel_parc_arb_others__h_total'),
                            h_poda_ = Avg('invrel_parc_arb_others__h_poda'),
                            area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb_others__dap') * F('invrel_parc_arb_others__dap'))) / 4) / 10000) 
                            / F('size_parcela'),
                             porc_remo = ((F('h_poda_') - relevamiento.h_last_poda) / 
                                          (F('h_total_') - relevamiento.h_last_poda)) * 100, 
                                          tam_parc =  Avg('size_parcela')
                            ) \
                .values()

                context.update({'data_resumen' : data_resumen})

                 #calculo lo mismo de arriba pero sin resumir
                res = ResumenRelevamientos.objects.get(relevamientos = relevamiento)
  
               
                context.update({'resumen_general' : res})

                damages = InventariosDamages.objects.all().order_by('pk')
                context.update({'damages' : damages})

                
                
                #data de los damages
                data_damages = ArbolesRelevamientoOthers.objects.filter(relevamientos_id = idrelevamiento) \
                 .values('parcela__number_parcela', 'damages_id') \
                 .annotate(res_damage = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                 .values('parcela__number_parcela', 'damages_id', 'damages__name', 'parcela__size_parcela', 'res_damage')\
                 .order_by('parcela__number_parcela')
                
                               
                
                
                #traigo las parcelas
                parcelas = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)

                #rearmo los damages
                data_damages_new = []
                
                for parc in parcelas:
                   
                    data_damages_present = []
                    #cargo los da;os presentes
                    #recorro los damages completos

                    for damag_ in damages:
                        present = False
                        #este tiene el dato agreado traido de la tabla
                        for dam_data in data_damages:

                            #primero consulto si estoy en la parcela

                            if parc.number_parcela == dam_data['parcela__number_parcela']:

                                  if dam_data['damages_id'] == damag_.pk:

                                    present = True
                                    data_damages_present.append({dam_data['damages_id'] : dam_data['res_damage']})
                        
                        #no encontro ese damages, entonces lo cargo vacio
                        if present == False:
                            data_damages_present.append({damag_.pk : ''})
                    
                    #cargo la parcela tmb
                    data_damages_new.append({
                        'parcela' : parc.number_parcela,
                        'damages' : data_damages_present
                    })

              
                context.update({'data_damages' : data_damages_new})
                
               

                observaciones = InventariosObservaciones.objects.all().order_by('pk')
                context.update({'observaciones' : observaciones})
                 #data de los damages

                data_calidad = ArbolesRelevamientoOthers.objects.filter(relevamientos_id = idrelevamiento) \
                 .values('parcela__number_parcela', 'observaciones_id') \
                 .annotate(res_calidad = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                 .values('parcela__number_parcela', 'observaciones_id', 'observaciones__name', 'parcela__size_parcela', 'res_calidad')\
                 .order_by('parcela__number_parcela')
                
             
                data_calidad_new = []

               


                for parc in parcelas:
                   
                    data_obs_present = []
                    #cargo los da;os presentes
                    #recorro los damages completos

                    for obs_ in observaciones:
                        present = False
                        #este tiene el dato agreado traido de la tabla
                        for cal_data in data_calidad:

                            #primero consulto si estoy en la parcela

                            if parc.number_parcela == cal_data['parcela__number_parcela']:

                                  if cal_data['observaciones_id'] == obs_.pk:

                                    present = True
                                    data_obs_present.append({cal_data['observaciones_id'] : cal_data['res_calidad']})
                        
                        #no encontro ese damages, entonces lo cargo vacio
                        if present == False:
                            data_obs_present.append({obs_.pk : ''})
                    
                    #cargo la parcela tmb
                    data_calidad_new.append({
                        'parcela' : parc.number_parcela,
                        'observaciones' : data_obs_present
                    })

              
                context.update({'data_observaciones' : data_calidad_new})

                



                return render(request, 'configuration/inventarios/relevamientos/view_statistics.html', context)


    except Exception as e:
         messages.error(request, str(e))
         return redirect('inventarios-relevamientos-index')
    

@login_required
def deleteRelevamiento(request, idrelevamiento):
    try:
        obj = InventariosRelevamientos.objects.get(pk=idrelevamiento)
        obj.delete()
       
        messages.success(request, "El Relevamiento ha sido eliminado.")
        return redirect('inventarios-relevamientos-index')

    except Rodales.DoesNotExist:
        raise Http404("error")
    except Exception as e:
        raise Http404(str(e))
    

@login_required
def deleteRelevamientoPre(request, idrelevamiento):
    try:
        obj = InventariosRelevamientos.objects.get(pk=idrelevamiento)
        obj.delete()
       
        messages.success(request, "El Relevamiento ha sido eliminado.")
        return redirect('inventarios-relevamientos-index-rel-prexistentes')

    except Rodales.DoesNotExist:
        raise Http404("error")
    except Exception as e:
        raise Http404(str(e))
 

@login_required
def editArbolesRelevamientoPoda(request, idrelevamiento, idarbol):

    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    

    try:
        context.update({'idrelevamiento' : idrelevamiento})
      
        
        arbolpoda = ArbolesRelevamientoPoda.objects.get(pk = idarbol)
        context.update({'arbolpoda' : arbolpoda})

        #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})

        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        #obtengo los datos
        
        if request.method == 'POST':
            
            dap = request.POST.get('dap')
            dmsm = request.POST.get('dmsm')
            h_poda = request.POST.get('h_poda')
            h_total = request.POST.get('h_total')
            is_poda = request.POST.get('is_poda')
            damages_ = request.POST.get('select-damage')
            observacion = request.POST.get('select-observacion')

            dap_ = dap if dap != '' else None
            dmsm_ = dmsm if dmsm != '' else None
            h_poda_ = h_poda if h_poda != '' else None
            h_total_ = h_total if h_total != '' else None

            arbolpoda.dap = dap_
            arbolpoda.dmsm = dmsm_
            arbolpoda.h_poda = h_poda_
            arbolpoda.h_total = h_total_
            arbolpoda.is_poda = is_poda
            arbolpoda.damages_id = damages_
            arbolpoda.observaciones_id = observacion

            try:

                with transaction.atomic():
            
                    arbolpoda.save()

                    #agregue los arboes, entonces cargo las datas de resumen

                    result = calculate_resumen_poda(idrelevamiento)

                    if result == False:
                        raise
               
                messages.success(request, "El Arbol se ha editado con éxito")

                return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))

            except Exception as e:
                messages.error(request, str(e))
                return render(request, 'configuration/inventarios/relevamientos/edit_trees_poda.html', context)





        return render(request, 'configuration/inventarios/relevamientos/edit_trees_poda.html', context)

    except ArbolesRelevamientoPoda.DoesNotExist as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
       

    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
@login_required
def editArbolesRelevamientoOthers(request, idrelevamiento, idarbol):

    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    

    try:
        context.update({'idrelevamiento' : idrelevamiento})
      
        
        arbolpoda = ArbolesRelevamientoOthers.objects.get(pk = idarbol)
        context.update({'arbolpoda' : arbolpoda})

        #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})

        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        #obtengo los datos
        
        if request.method == 'POST':
            
            dap = request.POST.get('dap')
            h_poda = request.POST.get('h_poda')
            h_total = request.POST.get('h_total')

            damages_ = request.POST.get('select-damage')
            observacion = request.POST.get('select-observacion')

            dap_ = dap if dap != '' else None
         
            h_poda_ = h_poda if h_poda != '' else None
            h_total_ = h_total if h_total != '' else None

            arbolpoda.dap = dap_

            arbolpoda.h_poda = h_poda_
            arbolpoda.h_total = h_total_
            arbolpoda.damages_id = damages_
            arbolpoda.observaciones_id = observacion

            try:
                with transaction.atomic():
            
                    arbolpoda.save()

                    #agregue los arboes, entonces cargo las datas de resumen

                    result = calculate_resumen_others(idrelevamiento)
                  

                    if result == False:
                        raise
               
                messages.success(request, "El Arbol se ha editado con éxito")

                return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))

            except Exception as e:
                messages.error(request, str(e))
                return render(request, 'configuration/inventarios/relevamientos/edit_trees_poda_others.html', context)





        return render(request, 'configuration/inventarios/relevamientos/edit_trees_poda_others.html', context)

    except ArbolesRelevamientoPoda.DoesNotExist as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
       

    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    

@login_required
def deleteArbolPoda(request, idarbol):

    try:

        with transaction.atomic():
                
            obj = ArbolesRelevamientoPoda.objects.get(pk=idarbol)
            res = obj.delete()
            

            #agregue los arboes, entonces cargo las datas de resumen

            result = calculate_resumen_poda(obj.relevamientos_id)

            if result == False:
                raise
      

        if res[0] == 1:
            #resto a la parcela una unidad de arbol
            #obj.parcela.number_trees = obj.parcela.number_trees - 1
            #obj.parcela.save()
       
            messages.success(request, "El Arbol ha sido eliminado.")
        else: 
            messages.error(request, "El Arbol no se ha eliminado. Intente nuevamente")
        
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[obj.relevamientos.pk]))

    except Rodales.DoesNotExist:
        raise Http404("error")
    except Exception as e:
        raise Http404(str(e))
    


@login_required
def deleteArbolOthers(request, idarbol):

    try:
        obj = ArbolesRelevamientoOthers.objects.get(pk=idarbol)
        res = obj.delete()

        if res[0] == 1:
            #resto a la parcela una unidad de arbol
            #obj.parcela.number_trees = obj.parcela.number_trees - 1
            #obj.parcela.save()
       
            messages.success(request, "El Arbol ha sido eliminado.")
        else: 
            messages.error(request, "El Arbol no se ha eliminado. Intente nuevamente")
        
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[obj.relevamientos.pk]))

    except Rodales.DoesNotExist:
        raise Http404("error")
    except Exception as e:
        raise Http404(str(e))

@login_required
def addParcela(request, idrelevamiento):

    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    context.update({'idrelevamiento' : idrelevamiento})

    inventario_ent = InventariosRelevamientos.objects.get(pk = idrelevamiento)

    #traigo el maxnumber de la tabla

    max_number = InventariosParcelasgis.objects.filter(relevamiento = idrelevamiento).aggregate(max = Max('number_parcela'))
    context.update({'number_parcela' : max_number['max'] + 1})

    context.update({'types_parcels' : types_parcels})

    if request.method == 'POST':

        size_parcela = request.POST.get('size_parcela')
        type = request.POST.get('tipo')
        number_trees = request.POST.get('number_trees')

        result = InventariosParcelasgis.objects.create(size_parcela = size_parcela, type = type, number_trees = number_trees, 
                                                       rodales = inventario_ent.rodales, relevamiento = inventario_ent, 
                                                       number_parcela =  max_number['max'] + 1)


        messages.success(request, "El Relevamiento ha sido agregado con éxito!.")
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
        
    return render(request, 'configuration/inventarios/relevamientos/add_parcels.html', context)

@login_required
def addArbolesByParcelaPoda(request, idrelevamiento, idparcela):
    
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        
        parcela = InventariosParcelasgis.objects.get(pk = idparcela)

        #consulto la cantidad de arboles y la cantidad declarada
        cant_arb = len(parcela.invrel_parc_arb.all())

        #numero de arboles a crear
        num_trees = parcela.number_trees - cant_arb

        #si es 0 vuelvo al inicio con un error
        if num_trees <= 0:
            messages.error(request, str('Ya se han cargado todos los arboles a la Parcela'))
            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))


        #traigo el max
        max_number = ArbolesRelevamientoPoda.objects.filter(parcela = parcela).aggregate(max = Max('number_tree'))['max']

        context.update({'num_trees' : num_trees})
        context.update({'parc' : parcela})
        context.update({'idrelevamiento' : idrelevamiento})

         #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})

        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        if request.method == 'POST':
            parcela_number = request.POST.getlist('parcela')
            numbers = request.POST.getlist('number_tree')
            dap = request.POST.getlist('dap')
            dmsm = request.POST.getlist('dmsm')
            h_poda = request.POST.getlist('h_poda')
            h_total = request.POST.getlist('h_total')
            is_poda = request.POST.getlist('is_poda')
            damages = request.POST.getlist('select-damage')
            observacion = request.POST.getlist('select-observacion')


            


          


            lista_objetos = []
            for n in range(num_trees):


                  #convierto
                dap_ = dap[n] if dap[n] != '' else None
                dmsm_ = dmsm[n] if dmsm[n] != '' else None
                h_poda_ = h_poda[n] if h_poda[n] != '' else None
                h_total_ = h_total[n] if h_total[n] != '' else None
                is_poda_ = is_poda[n] if is_poda[n] != '' else False

              

                num_t = int(numbers[n]) + int(max_number if max_number != None else 0)

                
                lista_objetos.append(ArbolesRelevamientoPoda(number_tree = num_t , dap = dap_, 
                                                             dmsm = dmsm_, h_poda = h_poda_, 
                                                             h_total = h_total_, 
                                                is_poda = is_poda_, damages_id = damages[n], 
                                                parcela_id = parcela_number[n], 
                                                observaciones_id = observacion[n],
                                                relevamientos_id = idrelevamiento ))


           

            with transaction.atomic():
            
                ArbolesRelevamientoPoda.objects.bulk_create(lista_objetos)

                #agregue los arboes, entonces cargo las datas de resumen

                result = calculate_resumen_poda(idrelevamiento)

                if result == False:
                    raise

            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
       


    except InventariosParcelasgis.DoesNotExist as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    #add_trees_by_parcels.html
    return render(request, 'configuration/inventarios/relevamientos/add_trees_by_parcels.html', context)




    
@login_required
def addArbolesByParcelaOthers(request, idrelevamiento, idparcela):
    
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        
        parcela = InventariosParcelasgis.objects.get(pk = idparcela)

        #consulto la cantidad de arboles y la cantidad declarada
        cant_arb = len(parcela.invrel_parc_arb_others.all())

        #numero de arboles a crear
        num_trees = parcela.number_trees - cant_arb

        #si es 0 vuelvo al inicio con un error
        if num_trees <= 0:
            messages.error(request, str('Ya se han cargado todos los arboles a la Parcela'))
            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))


        #traigo el max
        max_number = ArbolesRelevamientoOthers.objects.filter(parcela = parcela).aggregate(max = Max('number_tree'))['max']

        context.update({'num_trees' : num_trees})
        context.update({'parc' : parcela})
        context.update({'idrelevamiento' : idrelevamiento})

         #traigo los daños
        damages = InventariosDamages.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'damages' : damages})

        observaciones = InventariosObservaciones.objects.all().values_list('pk', 'name').order_by('name')
        context.update({'observaciones' : observaciones})

        if request.method == 'POST':
          
            numbers = request.POST.getlist('number_tree')
            dap = request.POST.getlist('dap')
            h_poda = request.POST.getlist('h_poda')
            h_total = request.POST.getlist('h_total')
            damages = request.POST.getlist('select-damage')
            observacion = request.POST.getlist('select-observacion')



            lista_objetos = []
            for n in range(num_trees):

                num_t = int(numbers[n]) + int(max_number)

                dap_ = dap[n] if dap[n] != '' else None
                h_poda_ = h_poda[n] if h_poda[n] != '' else None
                h_total_ = h_total[n] if h_total[n] != '' else None
                
                lista_objetos.append(ArbolesRelevamientoOthers(number_tree = num_t , 
                                                               dap = dap_, h_poda = h_poda_, 
                                                               h_total = h_total_, 
                                                damages_id = damages[n], parcela = parcela, 
                                                observaciones_id = observacion[n],
                                                relevamientos_id = idrelevamiento ))


            ArbolesRelevamientoOthers.objects.bulk_create(lista_objetos)

            messages.success(request, "Los Arboles han sido agregados con éxito!.")
       

            return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
       


    except InventariosParcelasgis.DoesNotExist as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    #add_trees_by_parcels.html
    return render(request, 'configuration/inventarios/relevamientos/add_trees_by_parcels_others.html', context)

@login_required
def editParcela(request, idrelevamiento, idparcela):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        parcela = InventariosParcelasgis.objects.get(pk = idparcela)
        context.update({'parcela' : parcela})

        context.update({'idrelevamiento' : idrelevamiento})
        
        context.update({'types_parcels' : types_parcels})


        if request.method == 'POST':

            size_parcela = request.POST.get('size_parcela')
            type = request.POST.get('tipo')
            number_trees = request.POST.get('number_trees')

            parcela.size_parcela = size_parcela
            parcela.type = type
            parcela.number_trees = number_trees


            try:
                parcela.save()
                messages.success(request, "La Parcela se ha editado con éxito")

                return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))

            except Exception as e:
                messages.error(request, str(e))
                return render(request, 'configuration/inventarios/relevamientos/edit_parcels.html', context)


    except InventariosParcelasgis.DoesNotExist as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse("inventarios-view-arboles", args=[idrelevamiento]))
    
    return render(request, 'configuration/inventarios/relevamientos/edit_parcels.html', context)


def printStatistics(request, idrelevamiento):
    context = {}
    try:
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)
        context.update({'relevamiento' : relevamiento})

        plantacion = Plantaciones.objects.filter(rodales_id = relevamiento.rodales_id)

        context.update({'plantacion' : plantacion})

        rodales_gis = Rodalesgis.objects.filter(rodales_id = relevamiento.rodales_id)
        
        context.update({'rodales_gis' : rodales_gis})

        intensity = 0
        
        if relevamiento.type == 'Preexistente':
            res = ResumenRelevamientos.objects.get(relevamientos = relevamiento)
            context.update({'resumen_general' : res})
            return render(request, 'configuration/inventarios/relevamientos/view_statistics_print.html', context)
        
        else:

            if relevamiento.categorias:
                #no esta vacio entonces proceso
                if relevamiento.categorias.name == 'Poda':
                    
                    
                    arboles = relevamiento.invrel_arb_poda.all()
                    context.update({'arboles' : arboles})
                    context.update({'cant_arboles' : len(arboles)})

                    #cargo los datos de damages
                    damages = InventariosDamages.objects.all()
                    context.update({'damages' : damages})

                    observaciones = InventariosObservaciones.objects.all()
                    context.update({'observaciones' : observaciones})

                
                    data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento) \
                    .values('pk') \
                    .annotate(
                                n_arb_ha = Count('invrel_parc_arb') * 10000 / F('size_parcela'), #CALCULAR NUMERO DE ARB POR HA
                                arboles_podados = Count('invrel_parc_arb', 
                                                        filter=Q(invrel_parc_arb__is_poda = True)) * 10000 / F('size_parcela'),
                                arb_h_deseada = Count('invrel_parc_arb', 
                                                    filter=Q(invrel_parc_arb__h_poda__gte = relevamiento.h_deseada)) * 10000 / F('size_parcela'), 
                                arb_no_podados = Count('invrel_parc_arb',  filter=Q(invrel_parc_arb__is_poda = False)) * 10000 / F('size_parcela'),
                                dap_ = Avg('invrel_parc_arb__dap'),
                                h_total_ = Avg('invrel_parc_arb__h_total', filter=Q( invrel_parc_arb__h_total__gt = Value('0'))),
                                h_poda_ = Avg('invrel_parc_arb__h_poda', filter=Q( invrel_parc_arb__h_poda__gt = Value('0'))),
                                dmsm_ = Avg('invrel_parc_arb__dmsm', filter=Q( invrel_parc_arb__dmsm__gt = Value('0'))),
                                area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb__dap') * F('invrel_parc_arb__dap'))) / 4) / 10000) 
                                / F('size_parcela'),
                                porc_remo = ((F('h_poda_') - relevamiento.h_last_poda) / 
                                            (F('h_total_') - relevamiento.h_last_poda)) * 100, 
                                            tam_parc =  Avg('size_parcela')
                                ) \
                    .values()
                    
                    '''  arb_h_deseada = Count('invrel_parc_arb', filter=Q(invrel_parc_arb__h_total__gte = Value(relevamiento.h_deseada), 
                                                                    invrel_parc_arb__h_total__lt = (relevamiento.h_deseada + 3))) * 10000 / F('size_parcela'), 
                            '''
                
                #calculo lo mismo de arriba pero sin resumir

                    res = calculate_general_resumen_poda(data_resumen=data_resumen)
                
                    context.update({'resumen_general' : res[0]})
                
                    damages = InventariosDamages.objects.all().order_by('pk')
                    context.update({'damages' : damages})
                    
                    #data de los damages
                    data_damages = ArbolesRelevamientoPoda.objects.filter(relevamientos_id = idrelevamiento) \
                    .values('parcela__number_parcela', 'damages_id') \
                    .annotate(res_damage = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                    .values('parcela__number_parcela', 'damages_id', 'damages__name', 'parcela__size_parcela', 'res_damage')\
                    .order_by('parcela__number_parcela')
                    

                    
                    #traigo las parcelas
                    parcelas = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento)

                    #rearmo los damages
                    data_damages_new = []
                    
                    for parc in parcelas:
                    
                        data_damages_present = []
                        #cargo los da;os presentes
                        #recorro los damages completos

                        for damag_ in damages:
                            present = False
                            #este tiene el dato agreado traido de la tabla
                            for dam_data in data_damages:

                                #primero consulto si estoy en la parcela

                                if parc.number_parcela == dam_data['parcela__number_parcela']:

                                    if dam_data['damages_id'] == damag_.pk:

                                        present = True
                                        data_damages_present.append({dam_data['damages_id'] : dam_data['res_damage']})
                            
                            #no encontro ese damages, entonces lo cargo vacio
                            if present == False:
                                data_damages_present.append({damag_.pk : ''})
                        
                        #cargo la parcela tmb
                        data_damages_new.append({
                            'parcela' : parc.number_parcela,
                            'damages' : data_damages_present
                        })

                
                    context.update({'data_damages' : data_damages_new})


                    observaciones = InventariosObservaciones.objects.all().order_by('pk')
                    context.update({'observaciones' : observaciones})
                    #data de los damages
                    
                    data_calidad = ArbolesRelevamientoPoda.objects.filter(relevamientos_id = idrelevamiento) \
                    .values('parcela__number_parcela', 'observaciones_id') \
                    .annotate(res_calidad = (Count('pk')  * 10000 / F('parcela__size_parcela')))\
                    .values('parcela__number_parcela', 'observaciones_id', 'observaciones__name', 'parcela__size_parcela', 'res_calidad')\
                    .order_by('parcela__number_parcela')
                    

                    data_calidad_new = []


                    for parc in parcelas:
                    
                        data_obs_present = []
                        #cargo los da;os presentes
                        #recorro los damages completos

                        for obs_ in observaciones:
                            present = False
                            #este tiene el dato agreado traido de la tabla
                            for cal_data in data_calidad:

                                #primero consulto si estoy en la parcela

                                if parc.number_parcela == cal_data['parcela__number_parcela']:

                                    if cal_data['observaciones_id'] == obs_.pk:

                                        present = True
                                        data_obs_present.append({cal_data['observaciones_id'] : cal_data['res_calidad']})
                            
                            #no encontro ese damages, entonces lo cargo vacio
                            if present == False:
                                data_obs_present.append({obs_.pk : ''})
                        
                        #cargo la parcela tmb
                        data_calidad_new.append({
                            'parcela' : parc.number_parcela,
                            'observaciones' : data_obs_present
                        })

                
                    context.update({'data_observaciones' : data_calidad_new})

                    

                    context.update({'data_resumen' : data_resumen})

                    return render(request, 'configuration/inventarios/relevamientos/view_statistics_print.html', context)


                else:
                    #aca traigo los datos para others arboles
                    arboles = relevamiento.invrel_arb_others.all()
                    context.update({'arboles' : arboles})
                    context.update({'cant_arboles' : len(arboles)})

                    #cargo los datos de damages
                    damages = InventariosDamages.objects.all()
                    context.update({'damages' : damages})

                    observaciones = InventariosObservaciones.objects.all()
                    context.update({'observaciones' : observaciones})

                    #CUANDO VIENEN VALORES VACIOS NO CONSIDERAR EN LOS CALCULOS DE LOS ARBOELS

                
                    data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento) \
                    .values('pk') \
                    .annotate(
                                n_arb_ha = Count('invrel_parc_arb_others'),
                                arb_h_deseada = Count('invrel_parc_arb_others', 
                                                    filter=Q(invrel_parc_arb_others__h_total__gte = relevamiento.h_deseada, 
                                                                    invrel_parc_arb_others__h_total__lt = (relevamiento.h_deseada + 3))), 
                                dap_ = Avg('invrel_parc_arb_others__dap'),
                                h_total_ = Avg('invrel_parc_arb_others__h_total'),
                                h_poda_ = Avg('invrel_parc_arb_others__h_poda'),
                                area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb_others__dap') * F('invrel_parc_arb_others__dap'))) / 4) / 10000) 
                                / F('size_parcela'),
                                porc_remo = ((F('h_poda_') - relevamiento.h_last_poda) / 
                                            (F('h_total_') - relevamiento.h_last_poda)) * 100, 
                                            tam_parc =  Avg('size_parcela')
                                ) \
                    .values()

                    context.update({'data_resumen' : data_resumen})

                    return render(request, 'configuration/inventarios/relevamientos/view_statistics_print.html', context) 
                



                return render(request, 'configuration/inventarios/relevamientos/view_statistics.html', context)


    except Exception as e:
         messages.error(request, str(e))
         return redirect('inventarios-relevamientos-index')


def chooseOptionInventarioPreexistente(request, idrelevamiento):

    try:
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

        #controlo si tengo cargado el resumen o sino envio a la carga
        cant_res = relevamiento.invrel_resumen.all()

        if cant_res > 0:
            #direcciono a ver
            pass
        
        else:
            #direcciono a cargar
            pass

    except InventariosRelevamientos.DoesNotExist as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')

    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')


def addResumenRelevamientoPreexistente(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Crea un Relevamiento de Inventario'}
    

    context.update({'idrelevamiento' : idrelevamiento})
    try:

        #taigo el relevamiento
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

        
        if request.method == 'POST':

            if relevamiento.categorias.name == 'Poda':

                #controlo los ingresos  vacios
                arb_ha = request.POST.get('arb_ha') if request.POST.get('arb_ha') != '' else None
                arb_podados = request.POST.get('arb_podados') if request.POST.get('arb_podados') != '' else None
                arb_h_deseada = request.POST.get('arb_h_deseada') if request.POST.get('arb_h_deseada') != '' else None
                arb_no_podados = request.POST.get('arb_no_podados') if request.POST.get('arb_no_podados') != '' else None
                dap = request.POST.get('dap') if request.POST.get('dap') != '' else None
                h_total = request.POST.get('h_total') if request.POST.get('h_total') != '' else None
                h_poda = request.POST.get('h_poda') if request.POST.get('h_poda') != '' else None
                dmsm = request.POST.get('dmsm') if request.POST.get('dmsm') != '' else None
                area_basal = request.POST.get('area_basal') if request.POST.get('area_basal') != '' else None
                copa_rem = request.POST.get('copa_rem') if request.POST.get('copa_rem') != '' else None

                user_entity = Users.objects.get(pk=request.user.pk)

                result = ResumenRelevamientos.objects.create(arb_ha = arb_ha, arb_podados = arb_podados, arb_h_deseada = arb_h_deseada, 
                                                            arb_no_podados = arb_no_podados, dap = dap, h_total = h_total, 
                                                            h_poda = h_poda, dmsm = dmsm, area_basal = area_basal, copa_rem = copa_rem,
                                                            user = user_entity, relevamientos_id = idrelevamiento)


            
            
                messages.success(request, "El Relevamiento ha sido cargado con éxito!.")
                #puedo direccionar a ver el resumen
                return redirect('inventarios-relevamientos-index')
            
            else:

                #controlo los ingresos  vacios
                arb_ha = request.POST.get('arb_ha') if request.POST.get('arb_ha') != '' else None
              
                dap = request.POST.get('dap') if request.POST.get('dap') != '' else None
                h_total = request.POST.get('h_total') if request.POST.get('h_total') != '' else None
                h_poda = request.POST.get('h_poda') if request.POST.get('h_poda') != '' else None
                
                area_basal = request.POST.get('area_basal') if request.POST.get('area_basal') != '' else None
             

                user_entity = Users.objects.get(pk=request.user.pk)

                result = ResumenRelevamientos.objects.create(arb_ha = arb_ha, 
                                                           dap = dap, h_total = h_total, 
                                                            h_poda = h_poda, area_basal = area_basal,
                                                            user = user_entity, relevamientos_id = idrelevamiento)


            
            
                messages.success(request, "El Relevamiento ha sido cargado con éxito!.")
                #puedo direccionar a ver el resumen
                return redirect('inventarios-relevamientos-index')

        

        if relevamiento.categorias.name == 'Poda':

        
            return render(request, 'configuration/inventarios/relevamientos/add_resumen_relevamiento_poda.html', context)
        else:
            return render(request, 'configuration/inventarios/relevamientos/add_resumen_relevamiento_others.html', context)


    except Exception as e:
            messages.error(request, str(e))
            return redirect('inventarios-relevamientos-index')
  

    
    

def viewResumenRelevamientoPreexistente(request, idrelevamiento):
    context = {
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Carga Árboles en un Relevamiento de Inventario'}
    
    try:
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)
        context.update({'relevamiento' : relevamiento})
       

        relevamiento_data = InventariosRelevamientos.objects.filter(pk = idrelevamiento).values('invrel_resumen__tam_parcela')
        context.update({'relevamiento_data' : relevamiento_data[0]})

        plantacion = Plantaciones.objects.filter(rodales_id = relevamiento.rodales_id)

        context.update({'plantacion' : plantacion})

        rodales_gis = Rodalesgis.objects.filter(rodales_id = relevamiento.rodales_id)
        
        context.update({'rodales_gis' : rodales_gis})

        resumen_general = ResumenRelevamientos.objects.get(relevamientos = relevamiento)

        print(relevamiento)
        print(resumen_general)
       
        context.update({'resumen_general' : resumen_general})
      


    except Exception as e:
        messages.error(request, str(e))

    return render(request, 'configuration/inventarios/relevamientos/view_statistics.html', context)



@login_required
def viewRelevamientosByRodal(request, idrodal):
    
    relevamientos = InventariosRelevamientos.objects.filter(rodales_id = idrodal) \
        .values(
        'pk', 'number', 'type', 'h_deseada', 'h_last_poda', 'fecha', 'created', 
        'is_finish', 'user__pk', 'user__first_name', 'user__last_name', 'user_relevador', 'user_relevador__first_name', 'user_relevador__last_name',
        'rodales_id', 'rodales__cod_sap', 'categorias_id', 'categorias__name', 'emsefor__name'
        ).annotate(cant_parcelas = Count('parc_invrel')) \
        .annotate(cant_arboles = Sum('parc_invrel__number_trees'), cant_res = Count('invrel_resumen', filter=Q(type = 'Preexistente'))).order_by('number')
    
  


    context = {'relevamientos' : relevamientos, 
               'category' : 'Relevamientos de Inventarios',
                'action' : 'Administración de los Relevamientos de Inventarios'}

    return render(request, 'configuration/inventarios/relevamientos/view_relevamientos_by_rodal.html', context)

@login_required
def addParcelaGis(request, idrelevamiento):

    
    context = {
            'category' : 'Rodales',
            'action' : 'Crea un nuevo Rodal'}
    
    relevamiento = InventariosRelevamientos.objects.get(pk=idrelevamiento)
    rodal = relevamiento.rodales


    rodaldata_ = Rodales.objects.filter(pk = rodal.pk)
    rodales_gis = serialize('geojson', Rodalesgis.objects.filter(rodales=rodal), geometry_field='geom_4326')

    context.update({'rodales_gis' : rodales_gis})
    context.update({'rodal' : rodal})

    rodaldata = serialize('json', rodaldata_)
    context.update({'rodaldata' : rodaldata})

    parcelas = list(InventariosParcelasgis.objects.filter(relevamiento = relevamiento).annotate(
    lat=Func("geom_4326", function="ST_Y", output_field=FloatField()),
    lon=Func("geom_4326", function="ST_X", output_field=FloatField()),).values('id', 'number_parcela', 'lat', 'lon')
    .order_by('number_parcela'))

  
    #parcelasdata = serialize('json', parcelas)
    context.update({'parcelasdata' : json.dumps({"data": parcelas})})

    parcelas_gis = serialize('geojson', InventariosParcelasgis.objects.filter(relevamiento = relevamiento), geometry_field='geom_4326')
    context.update({'parcelas_gis' : parcelas_gis})

    context.update({'idrelevamiento' : idrelevamiento})

    return render(request, 'configuration/inventarios/relevamientos/add_parcela_gis.html', context=context)


@login_required
def loadRelevamientosByExceIndex(request):
    
    context = {
            'category' : 'Inventarios',
            'action' : 'Cargar datos de Inventarios a partir de una Hoja de Calculo'}
    
    return render(request, 'configuration/inventarios/inventarios_load/index.html', context=context)




@login_required
def loadRelevamientosFiles(request, mode):
    context = {
            'category' : 'Inventarios',
            'action' : 'Cargar datos de Inventarios a partir de una Hoja de Calculo'}
    
    #mode puede ser 1 o 0, 0 tradicional y 1 poda

    context.update({
        'mode' : mode
    })

    if request.method == 'POST':

        if mode == 0:
            #proceso el inventario de PODA
            excel_file = request.FILES["excel_file"]
            data_poda = process_excel_poda(excel_file=excel_file)

            context.update({'data_poda' : data_poda})
            request.session["data_poda"] = data_poda


            return render(request, 'configuration/inventarios/inventarios_load/view_data_poda.html', context=context)



        elif mode == 1:

            excel_file = request.FILES["excel_file"]

                #controlo con un try catch algun error que pueda surgir y retorno false si algo sucede

            data_tradicional = process_excel_tradicional(excel_file=excel_file)

                #hago un redireccionamiento mostrando en una tabla los datos

            context.update({'data_tradicional' : data_tradicional})
            request.session["data_tradicional"] = data_tradicional

            return render(request, 'configuration/inventarios/inventarios_load/view_data_tradicional.html', context=context)

     
        
    
    return render(request, 'configuration/inventarios/inventarios_load/load_files.html', context=context)

@login_required
def saveRelevamientoFiles(request):

    try:
        #traigo el dato de la sesion
        data_tradicional = request.session.get("data_tradicional")
        #traigo la categoria
        categoria = InventariosCategories.objects.get(name = 'Tradicional')
       
        
        if categoria == None:
            raise
                    
        user_entity = Users.objects.get(pk=request.user.pk)

        #so esta variable para descargar los datos errados
        data_erroneos = []

        with transaction.atomic():
            #operaciones en la base de datos a completarse
            #primero creo el relevamiento y luego cargo el resumen
            #es crear un rele y a su vez el resumen
            resumen_rel_objects = []

            for data in data_tradicional:

                if data['is_present'] == True:

                    fecha = (str(data['fecha']) + str('-01-01')) if data['fecha'] != None else '0000-00-00'
                    rodal = data['rodal_pk'] if data['rodal_pk'] != None else None

                    number = len(InventariosRelevamientos.objects.filter(rodales_id = rodal).values()) + 1

                    #creo el relevamiento
                    relevamiento = InventariosRelevamientos.objects.create(
                        number = number, 
                        type = 'Preexistente',
                        fecha = fecha,
                        rodales_id = rodal,
                        user = user_entity,
                        user_relevador = user_entity, 
                        categorias = categoria, 
                        emsefor = None)
                    

                    #cargo el resumen inventario
                    num_arb = data['num_arb'] if data['num_arb'] != None else None
                    dap = data['dap'] if data['dap'] != None else None
                    h_total = data['h'] if data['h'] != None else None

                    area_basal = data['area_basal'] if data['area_basal'] != None else None
                    vmi_t = data['vmi_t'] if data['vmi_t'] != None else None
                    vmi_c = data['vmi_c'] if data['vmi_c'] != None else None

                    total = data['total'] if data['total'] != None else None
                    ima = data['ima'] if data['ima'] != None else None
                    
                    resumen_rel_objects.append(ResumenRelevamientos(
                        arb_ha = num_arb,
                        dap = dap,
                        h_total = h_total,
                        area_basal = area_basal,
                        vmi_t = vmi_t,
                        vmi_c = vmi_c,
                        total = total,
                        ima = ima,
                        relevamientos = relevamiento,
                        user = user_entity
                    ))
            
                else:

                    #agrego a una lista
                    data_erroneos.append(data)

            #almaceno con bulk
            ResumenRelevamientos.objects.bulk_create(resumen_rel_objects)

            #cargo en la session
            request.session["data_tradicional"] = None
            request.session["cant_rel_ok"] = len(resumen_rel_objects)
            request.session["data_erroneos"] = data_erroneos

            messages.success(request, "Los Relevamientos han sido cargado con éxito!.")
            #puedo direccionar a ver el resumen
            return redirect('inventarios-relevamientos-resumen-save-files-tradicional')




    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')

    

@login_required
def resumeSaveFileTradicional(request):
    context = {
            'category' : 'Inventarios',
            'action' : 'Cargar datos de Inventarios a partir de una Hoja de Calculo'}
    

    #traigo los datos de la session
    data_erroneos = request.session.get("data_erroneos")
    cant_rel_ok = request.session.get("cant_rel_ok")

    context.update({
        'data_erroneos' : data_erroneos
    })

    return render(request, 'configuration/inventarios/inventarios_load/view_data_tradicional_erroneos.html', context=context)



@login_required
def downloadExcelWithErrores(request):
    
    try:
        data_erroneos = request.session.get("data_erroneos")
       

        wb = openpyxl.Workbook()
        ws = wb.create_sheet(index=1, title="inventario_base")
        wb.active = wb['inventario_base']

        wb.remove_sheet(wb['Sheet'])

        headers = [
            'empresa', 'rodal', 'fecha', 'num_arb',	'dap', 'h',	'area_basal', 'vmi_t', 'vmi_c', 'vol_comercial', 
            'total', 'ima'
        ]

        data = [headers]

        for da in data_erroneos:
            data.append([
                da['empresa'], da['rodal'], da['fecha'], 
                da['num_arb'],	da['dap'], da['h'],	
                da['area_basal'], da['vmi_t'], da['vmi_c'], da['vol_comercial'], 
            da['total'], da['ima']
            ])

        for line in data:
            ws.append(line)
        #creo el libro



     
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=data_with_errores.xlsx'

        wb.save(response)
       
        return response

      
    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')



@login_required
def saveRelevamientosFilesPoda(request):

    try:
        #traigo el dato de la sesion
        data_tradicional = request.session.get("data_poda")
        #traigo la categoria
        categoria = InventariosCategories.objects.get(name = 'Poda')

        if categoria == None:
            raise

         #so esta variable para descargar los datos errados
        data_erroneos = []

        with transaction.atomic():
            #operaciones en la base de datos a completarse
            #primero creo el relevamiento y luego cargo el resumen
            #es crear un rele y a su vez el resumen
            resumen_rel_objects = []

            for data in data_tradicional:

                if data['is_present'] == True:

                    fecha = (str(data['fecha']) + str('-01-01')) if data['fecha'] != None else '0000-00-00'
                    rodal = data['rodal_pk'] if data['rodal_pk'] != None else None

                    

                    h_poda_ant = data['h_poda_ant'] if data['h_poda_ant'] != None else None
                    h_poda_ant = data['h_poda_ant'] if data['h_poda_ant'] != None else None
                    num_poda = data['num_poda'] if data['num_poda'] != None else None

                    arb_ha = data['arb_ha'] if data['arb_ha'] != None else None
                    arb_podados = data['arb_podados'] if data['arb_podados'] != None else None

                    arb_h_deseada = data['arb_h_deseada'] if data['arb_h_deseada'] != None else None
                    arb_no_podados = data['arb_no_podados'] if data['arb_no_podados'] != None else None
                    dap = data['dap'] if data['dap'] != None else None

                    h_total = data['h_total'] if data['h_total'] != None else None
                    h_poda = data['h_poda'] if data['h_poda'] != None else None

                    dmsm = data['dmsm'] if data['dmsm'] != None else None
                    area_basal = data['area_basal'] if data['area_basal'] != None else None
                    copa_rem = data['copa_rem'] if data['copa_rem'] != None else None
                    tam_parc = data['tam_parc'] if data['tam_parc'] != None else None

                     #creo el relevamiento
                    relevamiento = InventariosRelevamientos.objects.create(
                        number = num_poda, 
                        type = 'Preexistente',
                        h_last_poda = h_poda_ant,
                        fecha = fecha,
                        rodales_id = rodal,
                        user = request.user,
                        user_relevador = request.user, 
                        categorias = categoria, 
                        emsefor = None)
                    
                    #cargo el resumen inventario
                    #h_poda_ant	arb_ha	arb_podados	arb_h_deseada	arb_no_podados	dap	h_total	h_poda	dmsm	area_basal	copa_rem
       
                 

                    resumen_rel_objects.append(ResumenRelevamientos(
                        arb_ha = arb_ha,
                        arb_podados = arb_podados,
                        arb_h_deseada = arb_h_deseada,
                        arb_no_podados = arb_no_podados,
                        dap = dap,
                        h_total = h_total,
                        h_poda = h_poda,
                        dmsm = dmsm,
                        area_basal = area_basal,
                        copa_rem = copa_rem,
                        tam_parcela = tam_parc,
                        
                        relevamientos = relevamiento,
                        user = request.user, 
                    ))
            
                else:

                    #agrego a una lista
                    data_erroneos.append(data)

            
             #almaceno con bulk
            
            ResumenRelevamientos.objects.bulk_create(resumen_rel_objects)

            #cargo en la session
            request.session["data_poda"] = None
            request.session["cant_rel_ok"] = len(resumen_rel_objects)
            request.session["data_erroneos"] = data_erroneos

            messages.success(request, "Los Relevamientos han sido cargado con éxito!.")
            #puedo direccionar a ver el resumen
            return redirect('inventarios-relevamientos-resumen-save-files-poda')
    

    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')
    


@login_required
def resumeSaveFilePoda(request):
    context = {
            'category' : 'Inventarios',
            'action' : 'Cargar datos de Inventarios a partir de una Hoja de Calculo'}
    

    #traigo los datos de la session
    data_erroneos = request.session.get("data_erroneos")
    cant_rel_ok = request.session.get("cant_rel_ok")

    context.update({
        'data_erroneos' : data_erroneos
    })

    return render(request, 'configuration/inventarios/inventarios_load/view_data_poda_erroneos.html', context=context)


@login_required
def downloadExcelWithErroresPoda(request):
    
    try:
        data_erroneos = request.session.get("data_erroneos")
       

        wb = openpyxl.Workbook()
        ws = wb.create_sheet(index=1, title="inventario_base")
        wb.active = wb['inventario_base']

        wb.remove_sheet(wb['Sheet'])

        #rodal	fecha	num_poda	h_poda_ant	arb_ha	arb_podados	arb_h_deseada	arb_no_podados	dap	h_total	h_poda	dmsm	area_basal	copa_rem


        headers = [
            'rodal', 'fecha', 'num_poda', 'h_poda_ant',	'arb_ha', 'arb_podados',	
            'arb_h_deseada', 'arb_no_podados', 'dap', 'h_total', 
            'h_poda', 'dmsm', 'area_basal', 'copa_rem'
        ]

        data = [headers]

        '''
        'rodal' : row[0].value,
                    'fecha' : row[1].value,
                    'num_poda' : row[2].value,
                    'h_poda_ant' : row[3].value,
                    'arb_ha' : row[4].value,
                    'arb_podados' : row[5].value,
                    'arb_h_deseada' : row[6].value,
                    'arb_no_podados' : row[7].value,
                    'dap' : row[8].value,
                    'h_total' : row[9].value,
                    'h_poda' : row[10].value,
                    'dmsm' : row[11].value,
                    'area_basal' : row[12].value,
                    'copa_rem' : row[13].value,'''

        for da in data_erroneos:
            data.append([
                da['rodal'], da['fecha'], da['num_poda'], 
                da['h_poda_ant'],	da['arb_ha'], da['arb_podados'],	
                da['arb_h_deseada'], da['arb_no_podados'], da['dap'], da['h_total'], 
            da['h_poda'], da['dmsm'], da['area_basal'], da['copa_rem']
            ])

        for line in data:
            ws.append(line)
        #creo el libro



     
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=data_with_errores.xlsx'

        wb.save(response)
       
        return response

      
    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index')


@login_required
def indexRelevamientosPreexistentes(request):

    context = {
            'category' : 'Inventarios',
            'action' : 'Ver Relevamientos Preexistentes'}
    

    relevamientos = InventariosRelevamientos.objects.all() \
        .values(
        'pk', 'number', 'type', 'h_deseada', 'h_last_poda', 'fecha', 'created', 
        'is_finish', 'user__first_name', 'user__last_name', 'user_relevador', 'user_relevador__first_name', 'user_relevador__last_name',
        'rodales_id', 'rodales__cod_sap', 'categorias_id', 'categorias__name', 'emsefor__name',  'emsefor__pk',
            'invrel_resumen__is_load_to_intervencion', 'invrel_resumen__intervenciones__type', 
            'invrel_resumen__intervenciones__intervenciones_id', 'invrel_resumen__tam_parcela'
        ).annotate(cant_parcelas = Count('parc_invrel')) \
        .annotate(cant_arboles = Sum('parc_invrel__number_trees'), 
                  cant_res = Count('invrel_resumen', filter=Q(type = 'Preexistente'))).order_by('number')
    
  
    
    for rel in relevamientos:
            
            if rel['emsefor__name'] is not None:
                print(rel['pk'])
       

    context.update({'relevamientos' : relevamientos})

    return render(request, 'configuration/inventarios/relevamientos/view_relevamientos_preexistentes.html', context)

@login_required
def sendRelevamientoToIntervencion(request, idrelevamiento):

    context = {
            'category' : 'Inventarios',
            'action' : 'Ver Relevamientos Preexistentes'}
    
    try:
        
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

        rel_resumen = ResumenRelevamientos.objects.filter(relevamientos = relevamiento)[0]
       

        context.update({
            'relevamiento' : relevamiento,
            'rel_resumen' : rel_resumen
        })

        type_intervencion = INTERVENCIONES_DICT['Poda']

        emsefors = Emsefor.objects.values_list("emsefor_id", "name").order_by('name')
        context.update({'emsefors' : emsefors})

        type_intervecion = IntervencionesTypes.objects.values_list("intervencionestypes_id", "name").order_by('name')
        context.update({'type_intervecion' : type_intervecion})


        if request.method == 'POST':

            #recupero la informacion, pero utilizo rollback para almacenar
            name = request.POST.get("name")
            fecha = request.POST.get("fecha")
            superficie = request.POST.get("superficie")
            emsefor = request.POST.get("select-emsefor")
            inter_type_id = request.POST.get("select-inter-type")


            
            with transaction.atomic():
                
                intervencion = Intervenciones.objects.create(
                    rodales = relevamiento.rodales,
                    name = name,
                    fecha = fecha,
                    type = type_intervencion,
                    superficie = superficie,
                    users = request.user,
                    emsefors_id = emsefor,
                    intervenciones_types_id=inter_type_id
                )

                poda_intervencion = PodaIntervencion.objects.create(intervenciones=intervencion, 
                                                                    arb_podados=rel_resumen.arb_podados, 
                                                                    arb_no_podados=rel_resumen.arb_no_podados, 
                                                                        alt_deseada=rel_resumen.arb_h_deseada, 
                                                                        alt_poda=rel_resumen.h_poda, 
                                                                        dap=rel_resumen.dap, 
                                                                        altura=rel_resumen.h_total, 
                                                                        dmsm=rel_resumen.dmsm, 
                                                                        area_basal=rel_resumen.area_basal, 
                                                                        porc_removido=rel_resumen.copa_rem)
                
                #almaceno el id de la intervencion

                rel_resumen.is_load_to_intervencion = True
                rel_resumen.intervenciones = intervencion
                
                rel_resumen.save()

                relevamiento.emsefor_id = emsefor
                relevamiento.save()
                
                messages.success(request, "El Relevamiento de Poda ha sido enviado correctamente!.")
                return redirect('inventarios-relevamientos-index-rel-prexistentes')



        return render(request, 'configuration/inventarios/relevamientos_send/add.html', context)


    
    except InventariosRelevamientos.DoesNotExist as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index-rel-prexistentes')

    except Exception as e:
        messages.error(request, str(e))
        return redirect('inventarios-relevamientos-index-rel-prexistentes')




