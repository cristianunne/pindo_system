
import openpyxl
from rodales.models import Rodales
from inventario.models import InventariosRelevamientos, ResumenRelevamientos, InventariosParcelasgis

from django.db.models import Sum, F, Count, Q, Avg, Value, Max, Func, FloatField, ExpressionWrapper
from django.db.models.functions.comparison import NullIf
from django.db.models.functions import Concat

names_sheets_excel_tradicional = [
    'inventario_base'
]

names_sheets_excel_poda = [
    'inventario_poda'
]


def calculate_resumen_poda(idrelevamiento):

    try:

        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)
    
        data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = relevamiento.pk) \
                    .values('pk') \
                    .annotate(
                                n_arb_ha = Count('invrel_parc_arb') * 10000 / F('size_parcela'), #CALCULAR NUMERO DE ARB POR HA
                                arboles_podados = Count('invrel_parc_arb', 
                                                        filter=Q(invrel_parc_arb__is_poda = True)) * 10000 / F('size_parcela'),
                                arb_h_deseada = Count('invrel_parc_arb', 
                                                    filter=Q(invrel_parc_arb__h_poda__gte = relevamiento.h_deseada)) * 10000 / F('size_parcela'), 
                                arb_no_podados = Count('invrel_parc_arb',  filter=Q(invrel_parc_arb__is_poda = False)) * 10000 / F('size_parcela'),
                                dap_ = Avg('invrel_parc_arb__dap'),
                                h_total_ = Avg('invrel_parc_arb__h_total', filter=Q( invrel_parc_arb__h_total__gt = Value('0'))),
                                h_poda_ = Avg('invrel_parc_arb__h_poda', filter=Q( invrel_parc_arb__h_poda__gt = Value('0'))),
                                dmsm_ = Avg('invrel_parc_arb__dmsm', filter=Q( invrel_parc_arb__dmsm__gt = Value('0'))),
                                area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb__dap') * F('invrel_parc_arb__dap'))) / 4) / 10000) 
                                / F('size_parcela'),
                                porc_remo = ExpressionWrapper(((F('h_poda_') - relevamiento.h_last_poda) / 
                                          NullIf(F('h_total_') - relevamiento.h_last_poda, 0)) * 100, 
                                          output_field=FloatField()),
                                tam_parc =  Avg('size_parcela')
                                ).values()
        
        data_resumen_general = calculate_general_resumen_poda(data_resumen)

     

        res_relevamiento = ResumenRelevamientos.objects.get(relevamientos = relevamiento)

        if len(data_resumen) > 0:

            res_relevamiento.tam_parcela = data_resumen_general[0]['tam_parc']
            res_relevamiento.arb_ha = data_resumen_general[0]['n_arb_ha']
            res_relevamiento.arb_podados = data_resumen_general[0]['arboles_podados']
            res_relevamiento.arb_h_deseada = data_resumen_general[0]['arb_h_deseada']
            res_relevamiento.arb_no_podados = data_resumen_general[0]['arb_no_podados']

            res_relevamiento.dap = data_resumen_general[0]['dap']
            res_relevamiento.h_total = data_resumen_general[0]['h_total']
            res_relevamiento.h_poda = data_resumen_general[0]['h_poda']
            res_relevamiento.dmsm = data_resumen_general[0]['dmsm']
            res_relevamiento.area_basal = data_resumen_general[0]['area_basal']

            res_relevamiento.copa_rem = data_resumen_general[0]['porc_rem']

            res_relevamiento.save()

            return True
        



    except ResumenRelevamientos.DoesNotExist as e:
        return False
    
    except Exception as e:
        return False

    #obtengo el resumen de la tabla
    



def calculate_general_resumen_poda(data_resumen):

    n_arb_ha = 0
    arboles_podados = 0
    arb_h_deseada = 0
    arb_no_podados = 0
    dap = 0
    h_total = 0
    h_poda = 0
    dmsm = 0
    area_basal = 0
    proc_rem = 0
    tam_parc = 0

    n_arb_ha_index = 0
    arboles_podados_index = 0
    arb_h_deseada_index = 0
    arb_no_podados_index = 0
    dap_index = 0
    h_total_index = 0
    h_poda_index = 0
    dmsm_index = 0
    area_basal_index = 0
    proc_rem_index = 0
    tam_parc_index = 0

    for data in data_resumen:

      
        
        if data['n_arb_ha'] is not None:
            if data['n_arb_ha'] > 0:
                n_arb_ha = n_arb_ha + data['n_arb_ha']
                n_arb_ha_index = n_arb_ha_index + 1

        if data['arboles_podados'] is not None:
            if data['arboles_podados'] > 0:
                arboles_podados = arboles_podados + data['arboles_podados']
                arboles_podados_index = arboles_podados_index + 1
        
        if data['arb_h_deseada'] is not None:
            if data['arb_h_deseada'] > 0:
                arb_h_deseada = arb_h_deseada + data['arb_h_deseada']
                arb_h_deseada_index = arb_h_deseada_index + 1
        
        if data['arb_no_podados'] is not None:
            if data['arb_no_podados'] > 0:
                arb_no_podados = arb_no_podados + data['arb_no_podados']
                arb_no_podados_index = arb_no_podados_index + 1
        
        if data['dap_'] is not None:
            if data['dap_'] > 0:
                dap = dap + data['dap_']
                dap_index = dap_index + 1
        
        if data['h_total_'] is not None:
            if data['h_total_'] > 0:
                h_total = h_total + data['h_total_']
                h_total_index = h_total_index + 1
        
        if data['h_poda_'] is not None:
            if data['h_poda_'] > 0:
                h_poda = h_poda + data['h_poda_']
                h_poda_index = h_poda_index + 1
        
        if data['dmsm_'] is not None:
            if data['dmsm_'] > 0:
                dmsm = dmsm + data['dmsm_']
                dmsm_index = dmsm_index + 1
        
        if data['area_basal'] is not None:
            if data['area_basal'] > 0:
                area_basal = area_basal + data['area_basal']
                area_basal_index = area_basal_index + 1
        
        if data['porc_remo'] is not None:
            if data['porc_remo'] > 0:
                proc_rem = proc_rem + data['porc_remo']
                proc_rem_index = proc_rem_index + 1
            
            else:
                proc_rem_index = proc_rem_index + 1

        
        if data['tam_parc'] is not None:
            if data['tam_parc'] > 0:
                tam_parc = tam_parc + data['tam_parc']
                tam_parc_index = tam_parc_index + 1
            else:
                tam_parc_index = tam_parc_index + 1


    
    #controlo la / por 0
    result = [{
        'n_arb_ha' : n_arb_ha / n_arb_ha_index if n_arb_ha_index > 0 else 0,
        'arboles_podados' : arboles_podados / arboles_podados_index if arboles_podados_index > 0 else 0,
        'arb_h_deseada' : arb_h_deseada / arb_h_deseada_index if arb_h_deseada_index > 0 else 0,
        'arb_no_podados': arb_no_podados / arb_no_podados_index if arb_no_podados_index > 0 else 0,
        'dap' : dap / dap_index if dap_index > 0 else 0,
        'h_total' : h_total / h_total_index if h_total_index > 0 else 0,
        'h_poda' : h_poda / h_poda_index if h_poda_index > 0 else 0,
        'dmsm' : dmsm / dmsm_index if dmsm_index > 0 else 0,
        'area_basal' : area_basal / area_basal_index if area_basal_index > 0 else 0,
        'porc_rem' : proc_rem / proc_rem_index if proc_rem_index > 0 else 0,
        'tam_parc' : tam_parc / tam_parc_index if tam_parc_index > 0 else 0,


    }]
    print(result)
    return result


def calculate_resumen_others(idrelevamiento):

    try:
        relevamiento = InventariosRelevamientos.objects.get(pk = idrelevamiento)

        data_resumen = InventariosParcelasgis.objects.filter(relevamiento_id = idrelevamiento) \
                .values('pk') \
                .annotate(
                            n_arb_ha = Count('invrel_parc_arb_others'),
                            arb_h_deseada = Count('invrel_parc_arb_others', 
                                                  filter=Q(invrel_parc_arb_others__h_total__gte = relevamiento.h_deseada, 
                                                                 invrel_parc_arb_others__h_total__lt = (relevamiento.h_deseada + 3))), 
                            dap_ = Avg('invrel_parc_arb_others__dap'),
                            h_total_ = Avg('invrel_parc_arb_others__h_total'),
                            h_poda_ = Avg('invrel_parc_arb_others__h_poda'),
                            area_basal = 10000 * (Sum((3.14 * (F('invrel_parc_arb_others__dap') * F('invrel_parc_arb_others__dap'))) / 4) / 10000) 
                            / F('size_parcela'),
                           
                            tam_parc =  Avg('size_parcela')
                            ) \
                .values()
        
       

        data_resumen_general = calculate_general_resumen_others(data_resumen)

     

        res_relevamiento = ResumenRelevamientos.objects.get(relevamientos = relevamiento)

        if len(data_resumen) > 0:

            res_relevamiento.tam_parcela = data_resumen_general[0]['tam_parc']
            res_relevamiento.arb_ha = data_resumen_general[0]['n_arb_ha']
            res_relevamiento.arb_h_deseada = data_resumen_general[0]['arb_h_deseada']

            res_relevamiento.dap = data_resumen_general[0]['dap']
            res_relevamiento.h_total = data_resumen_general[0]['h_total']
            res_relevamiento.h_poda = data_resumen_general[0]['h_poda']
            res_relevamiento.area_basal = data_resumen_general[0]['area_basal']


            res_relevamiento.save()

            return True

    except ResumenRelevamientos.DoesNotExist as e:
       
        return False
    
    except Exception as e:

        return False


def calculate_general_resumen_others(data_resumen):

    n_arb_ha = 0
    arb_h_deseada = 0
    dap = 0
    h_total = 0
    h_poda = 0
    area_basal = 0

    tam_parc = 0

    n_arb_ha_index = 0
    arb_h_deseada_index = 0
    dap_index = 0
    h_total_index = 0
    h_poda_index = 0
    area_basal_index = 0

    tam_parc_index = 0

    for data in data_resumen:

      
        
        if data['n_arb_ha'] is not None:
            if data['n_arb_ha'] > 0:
                n_arb_ha = n_arb_ha + data['n_arb_ha']
                n_arb_ha_index = n_arb_ha_index + 1

        
        if data['arb_h_deseada'] is not None:
            if data['arb_h_deseada'] > 0:
                arb_h_deseada = arb_h_deseada + data['arb_h_deseada']
                arb_h_deseada_index = arb_h_deseada_index + 1
        
        
        if data['dap_'] is not None:
            if data['dap_'] > 0:
                dap = dap + data['dap_']
                dap_index = dap_index + 1
        
        if data['h_total_'] is not None:
            if data['h_total_'] > 0:
                h_total = h_total + data['h_total_']
                h_total_index = h_total_index + 1
        
        if data['h_poda_'] is not None:
            if data['h_poda_'] > 0:
                h_poda = h_poda + data['h_poda_']
                h_poda_index = h_poda_index + 1
        
        
        if data['area_basal'] is not None:
            if data['area_basal'] > 0:
                area_basal = area_basal + data['area_basal']
                area_basal_index = area_basal_index + 1
        
        
        if data['tam_parc'] is not None:
            if data['tam_parc'] > 0:
                tam_parc = tam_parc + data['tam_parc']
                tam_parc_index = tam_parc_index + 1
            else:
                tam_parc_index = tam_parc_index + 1


    
    #controlo la / por 0
    result = [{
        'n_arb_ha' : n_arb_ha / n_arb_ha_index if n_arb_ha_index > 0 else 0,
        'arb_h_deseada' : arb_h_deseada / arb_h_deseada_index if arb_h_deseada_index > 0 else 0,
        'dap' : dap / dap_index if dap_index > 0 else 0,
        'h_total' : h_total / h_total_index if h_total_index > 0 else 0,
        'h_poda' : h_poda / h_poda_index if h_poda_index > 0 else 0,
        'area_basal' : area_basal / area_basal_index if area_basal_index > 0 else 0,
        'tam_parc' : tam_parc / tam_parc_index if tam_parc_index > 0 else 0,


    }]

    return result




def process_excel_tradicional(excel_file):
    wb = openpyxl.load_workbook(excel_file)
  
    worksheet = wb.get_sheet_by_name(names_sheets_excel_tradicional[0])
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row

    data_trad_list = []
    
    saps_rodales_idxs = []

    index = 0

    #empresa	rodal	fecha	num_arb	dap	h	area_basal	vmi_t	vmi_c	vol_comercial	total	ima

    for row in worksheet.iter_rows():
        
        if index > 0:

            #cargo todos los codigos SAPS
            saps_rodales_idxs.append(row[1].value)

            data_trad_list.append(
                {
                    'empresa' : row[0].value,
                    'rodal' : row[1].value,
                    'fecha' : row[2].value,
                    'num_arb' : row[3].value,
                    'dap' : row[4].value,
                    'h' : row[5].value,
                    'area_basal' : row[6].value,
                    'vmi_t' : row[7].value,
                    'vmi_c' : row[8].value,
                    'vol_comercial' : row[9].value,
                    'total' : row[10].value,
                    'ima' : row[11].value,

                }
            )
        else:
            index = index + 1

    rodales_present = get_rodales_by_sap(saps_rodales_idxs)
   

    for data in data_trad_list:

        is_present = False
        pk_rodal = None

        #recorro los rodales
        for rod in rodales_present:

            if data['rodal'] == rod.cod_sap:
                is_present = True
                pk_rodal = rod.pk

        data.update({'is_present' : is_present})
        data.update({'rodal_pk' : pk_rodal})
   
    
    return data_trad_list
 


def get_rodales_by_sap(idx_saps):
    
    rodales = Rodales.objects.filter(cod_sap__in = idx_saps)
    
    return rodales

def process_excel_poda(excel_file):

    wb = openpyxl.load_workbook(excel_file)
  
    worksheet = wb.get_sheet_by_name(names_sheets_excel_poda[0])
    excel_data = list()

    # iterating over the rows and
    # getting value from each cell in row

    data_trad_list = []
    
    saps_rodales_idxs = []

    index = 0

    for row in worksheet.iter_rows():
        
        if index > 0:

            #cargo todos los codigos SAPS
            saps_rodales_idxs.append(row[0].value)
            #h_poda_ant	arb_ha	arb_podados	arb_h_deseada	arb_no_podados	dap	h_total	h_poda	dmsm	area_basal	copa_rem
       

            data_trad_list.append(
                {
                    'rodal' : row[0].value,
                    'fecha' : row[1].value,
                    'num_poda' : row[2].value,
                    'h_poda_ant' : row[3].value,
                    'arb_ha' : row[4].value,
                    'arb_podados' : row[5].value,
                    'arb_h_deseada' : row[6].value,
                    'arb_no_podados' : row[7].value,
                    'dap' : row[8].value,
                    'h_total' : row[9].value,
                    'h_poda' : row[10].value,
                    'dmsm' : row[11].value,
                    'area_basal' : row[12].value,
                    'copa_rem' : row[13].value,
                    'tam_parc' : row[14].value,

                }
            )
        else:
            index = index + 1

    rodales_present = get_rodales_by_sap(saps_rodales_idxs)


    for data in data_trad_list:

        is_present = False
        pk_rodal = None

        #recorro los rodales
        for rod in rodales_present:

            if data['rodal'] == rod.cod_sap:
                is_present = True
                pk_rodal = rod.pk

        data.update({'is_present' : is_present})
        data.update({'rodal_pk' : pk_rodal})
   
    
    return data_trad_list